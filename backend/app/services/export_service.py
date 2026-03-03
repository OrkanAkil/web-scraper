import io
import csv
import json
from typing import List, Dict, Any
from openpyxl import Workbook


class ExportService:
    """Handles exporting scraped data to CSV, JSON, and Excel formats."""

    def to_csv(self, data: List[Dict[str, Any]], fields: List[str] = None) -> bytes:
        if not data:
            return b""

        if not fields:
            fields = list(data[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow(row)

        return output.getvalue().encode("utf-8-sig")

    def to_json(self, data: List[Dict[str, Any]]) -> bytes:
        return json.dumps(data, ensure_ascii=False, indent=2, default=str).encode("utf-8")

    def to_excel(self, data: List[Dict[str, Any]], fields: List[str] = None) -> bytes:
        if not data:
            return b""

        if not fields:
            fields = list(data[0].keys())

        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Data"

        # Header row with styling
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = cell.font.copy(bold=True)

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row_data.get(field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, 1):
            max_length = len(str(field))
            for row_idx in range(2, min(len(data) + 2, 100)):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


export_service = ExportService()
